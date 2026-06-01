"""
pdp_download.py — Download do PDP ENEVA N via PDPW/SINtegre
────────────────────────────────────────────────────────────
Uso:
    python pdp_download.py                   # D-1 (padrão)
    python pdp_download.py --data 2026-05-30
    python pdp_download.py --visible         # browser visível (debug)
    python pdp_download.py --agendar         # cria tarefa Windows 20:30

Saída:
    PDP_ENEVA_N_AAAA-MM-DD.csv   — dados estruturados (Data, Hora, Codigo, UGs, MW)
    PDP_ENEVA_N_AAAA-MM-DD.xlsx  — planilha convertida

Credenciais: variáveis de ambiente ONS_USUARIO / ONS_SENHA ou arquivo .env
"""

import argparse, csv, io, os, re, shutil, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from datetime import datetime, timedelta
from pathlib  import Path

try:
    from dotenv import load_dotenv; load_dotenv()
except ImportError:
    pass

from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout

# ─── Configuração ────────────────────────────────────────────────────────────

CONFIG = {
    "usuario":       os.getenv("ONS_USUARIO", ""),
    "senha":         os.getenv("ONS_SENHA",   ""),
    "empresa_label": "ENEVA N",
    "empresa_code":  "PX",          # código PDPW para ENEVA N (valor no select)
    "tipo_dado":     "4",           # 4 = Consistidos DESSEM
    "pasta_destino": r"G:\Meu Drive\OM_Parnaiba_Dashboard\dev",
    "pasta_temp":    str(Path.home() / "Downloads" / "pdp_temp"),
    "headless":      True,
    "timeout":       30000,
}

# URL base do relatório XLS — parâmetros fixos para ENEVA N
PLANILHA_URL_TPL = (
    "https://pdpw.ons.org.br/pdp/frmPlanilha.aspx"
    "?strDataPDP={data}"
    "&strCampo=sup"
    "&strEmpresa={code}|{label}"
    "&strTabela=despa"
    "&strBase=pdp"
    "&strAcesso="
)

# ─── Mapeamento código PDPW → UGs do dashboard ───────────────────────────────

MAPA_PDPW = {
    "PXUTP5": ["TV28"],
    "PXUTM3": ["TG51", "TG52", "TV58"],
    "PXUTM4": ["TG32", "TG31"],
    "PXUTM5": ["TG22", "TG21"],
    "PXUTP4": ["BAG011", "BAG021", "BAG031"],
    "PXUTNV": ["TG12", "UG10", "TV18"],
}
COLUNAS_INTERESSE = list(MAPA_PDPW.keys())

# ─── Helpers ─────────────────────────────────────────────────────────────────

def log(msg, nivel="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    p  = {"INFO": "i", "OK": "v", "ERRO": "X", "AVISO": "!"}
    print(f"[{ts}] [{p.get(nivel,'?')}] {msg}")


def validar_config():
    erros = []
    if not CONFIG["usuario"]: erros.append("ONS_USUARIO nao definido (configure .env)")
    if not CONFIG["senha"]:   erros.append("ONS_SENHA nao definida (configure .env)")
    Path(CONFIG["pasta_destino"]).mkdir(parents=True, exist_ok=True)
    Path(CONFIG["pasta_temp"]).mkdir(parents=True, exist_ok=True)
    if erros:
        for e in erros: log(e, "ERRO")
        sys.exit(1)


def limpar_temp():
    for pat in ("Relatorio*.xls*", "PDP*.xls*"):
        for f in Path(CONFIG["pasta_temp"]).glob(pat):
            try: f.unlink()
            except: pass

# ─── Download principal ───────────────────────────────────────────────────────

def baixar_pdp(data_iso: str) -> Path | None:
    data_val = data_iso.replace("-", "")        # "2026-05-30" → "20260530"
    url = PLANILHA_URL_TPL.format(
        data  = data_val,
        code  = CONFIG["empresa_code"],
        label = CONFIG["empresa_label"],
    )
    destino_base = Path(CONFIG["pasta_destino"]) / f"PDP_ENEVA_N_{data_iso}"
    limpar_temp()

    log(f"Iniciando download PDP ENEVA N · {data_iso}")
    log(f"URL: {url}")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=CONFIG["headless"],
            args=["--no-sandbox", "--disable-gpu"],
        )
        ctx  = browser.new_context(
            accept_downloads=True,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        )
        page = ctx.new_page()
        page.set_default_timeout(CONFIG["timeout"])

        try:
            # ── Login SINtegre ────────────────────────────────────────────
            log("Login SINtegre...")
            page.goto("https://sintegre.ons.org.br", wait_until="domcontentloaded", timeout=60000)
            page.fill("#username", CONFIG["usuario"])
            page.fill("#password", CONFIG["senha"])
            page.click("input[value='Entrar']")
            page.wait_for_load_state("domcontentloaded", timeout=60000)
            log("Login OK", "OK")

            # ── Download direto via frmPlanilha.aspx ──────────────────────
            # O servidor responde com Content-Disposition: attachment, então
            # page.goto nunca dispara "domcontentloaded" — ignoramos o timeout
            # do goto e deixamos expect_download capturar o arquivo.
            log("Solicitando planilha XLS...")
            with page.expect_download(timeout=90000) as dl_info:
                try:
                    page.goto(url, wait_until="commit", timeout=90000)
                except Exception:
                    pass  # esperado: download não carrega página
            dl = dl_info.value

            arq_temp = Path(CONFIG["pasta_temp"]) / (dl.suggested_filename or "Relatorio.xls")
            dl.save_as(str(arq_temp))
            log(f"XLS baixado: {arq_temp.name}", "OK")

            browser.close()

            # ── Converter XLS → XLSX ──────────────────────────────────────
            return _converter_para_xlsx(arq_temp, destino_base)

        except PwTimeout as e:
            log(f"Timeout: {e}", "ERRO")
            _salvar_screenshot(page, "erro_timeout")
            browser.close()
            return None
        except Exception as e:
            log(f"Erro: {e}", "ERRO")
            _salvar_screenshot(page, "erro_inesperado")
            browser.close()
            return None


def _salvar_screenshot(page, nome):
    try:
        p = Path(CONFIG["pasta_temp"]) / f"{nome}_{datetime.now():%H%M%S}.png"
        page.screenshot(path=str(p))
        log(f"Screenshot: {p}", "AVISO")
    except: pass

# ─── Conversão XLS → XLSX ────────────────────────────────────────────────────

def _converter_para_xlsx(arq_xls: Path, destino_base: Path) -> Path:
    """
    O PDPW gera um arquivo .xls que na verdade é TSV (tab-separado, Latin-1).
    Tenta abrir com xlrd (XLS binário real). Se falhar, trata como TSV e
    converte para XLSX com openpyxl.
    """
    destino_xlsx = destino_base.with_suffix(".xlsx")
    try:
        import xlrd, openpyxl
        log("Tentando xlrd (XLS binario)...")
        wb_old = xlrd.open_workbook(str(arq_xls))
        wb_new = openpyxl.Workbook()
        wb_new.remove(wb_new.active)
        for sname in wb_old.sheet_names():
            ws_o = wb_old.sheet_by_name(sname)
            ws_n = wb_new.create_sheet(title=sname)
            for r in range(ws_o.nrows):
                for c in range(ws_o.ncols):
                    ws_n.cell(r+1, c+1, ws_o.cell_value(r, c))
        wb_new.save(str(destino_xlsx))
        try: arq_xls.unlink()
        except: pass
        log(f"XLSX via xlrd: {destino_xlsx.name}", "OK")
        return destino_xlsx
    except Exception as e:
        log(f"xlrd falhou ({type(e).__name__}) — interpretando como TSV", "AVISO")

    # Formato TSV/HTML com extensão .xls (gerado pelo PDPW)
    try:
        import openpyxl
        texto = arq_xls.read_bytes().decode("latin-1")
        linhas = texto.splitlines()

        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = "PDP"

        for linha in linhas:
            cols = [c.strip() for c in linha.split("\t")]
            row_vals = []
            for c in cols:
                # tenta converter número com vírgula decimal
                try:
                    row_vals.append(float(c.replace(",", ".")))
                except (ValueError, AttributeError):
                    row_vals.append(c)
            ws_new.append(row_vals)

        wb_new.save(str(destino_xlsx))
        try: arq_xls.unlink()
        except: pass
        log(f"XLSX via TSV: {destino_xlsx.name}", "OK")
        return destino_xlsx
    except Exception as e2:
        log(f"Conversao TSV falhou ({e2}) — mantendo XLS original", "AVISO")
        dest = destino_base.with_suffix(".xls")
        shutil.copy2(str(arq_xls), str(dest))
        return dest

# ─── Pós-processamento ────────────────────────────────────────────────────────

def _ler_tsv_pdpw(arq: Path) -> tuple[list[str], list[list]]:
    """
    Lê o arquivo TSV gerado pelo PDPW (encoding Latin-1, decimal vírgula).
    Retorna (headers, linhas_de_dados).
    A 1ª linha de dados úteis é a linha que começa com 'Intervalo'.
    """
    texto  = arq.read_bytes().decode("latin-1")
    linhas = texto.splitlines()

    headers    = []
    dados      = []
    header_idx = None

    for i, linha in enumerate(linhas):
        cols = [c.strip() for c in linha.split("\t")]
        if cols and cols[0].upper() == "INTERVALO":
            headers    = [c.upper() for c in cols]
            header_idx = i
            continue
        if header_idx is None:
            continue  # ainda não chegamos no cabeçalho
        if not cols or not cols[0]:
            continue
        skip = ("TOTAL", "MEDIA", "MÉDIA", "")
        if cols[0].strip().upper() in skip:
            continue
        dados.append(cols)

    return headers, dados


def processar_pdp(arq: Path, data_iso: str) -> Path:
    """
    Processa o arquivo baixado do PDPW (TSV ou XLSX convertido).
    Gera CSV estruturado e exibe resumo.
    """
    log("Processando colunas de interesse...")

    # Determinar fonte de leitura: TSV (.xls original) ou XLSX convertido
    if arq.suffix.lower() in (".xls",):
        headers, dados = _ler_tsv_pdpw(arq)
    else:
        # XLSX convertido — tentar xlrd via TSV também (o XLSX foi gerado a partir do TSV)
        # mas se o arquivo é um XLSX legítimo, ler normalmente
        try:
            import openpyxl
            wb  = openpyxl.load_workbook(str(arq), data_only=True)
            ws  = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            # encontrar linha de cabeçalho
            header_idx = None
            for i, row in enumerate(all_rows):
                rv = [str(c).strip().upper() if c else "" for c in row]
                if "INTERVALO" in rv:
                    header_idx = i
                    headers = rv
                    break
            if header_idx is None:
                log("Cabecalho nao encontrado no XLSX", "AVISO")
                for i, r in enumerate(all_rows[:5]):
                    log(f"  Linha {i+1}: {[str(c)[:15] for c in (r or [])[:10]]}")
                return arq
            dados_raw = all_rows[header_idx+1:]
            dados = []
            for row in dados_raw:
                if not row or all(c is None for c in row): continue
                cols = [str(c).strip() if c is not None else "" for c in row]
                if not cols[0] or cols[0].upper() in ("TOTAL","MEDIA","MÉDIA"): continue
                dados.append(cols)
        except Exception as e:
            log(f"Erro ao ler XLSX ({e}) — tentando como TSV", "AVISO")
            headers, dados = _ler_tsv_pdpw(arq)

    if not headers:
        log("Nenhum cabecalho encontrado", "ERRO")
        return arq

    # Mapear colunas de interesse
    col_map = {}
    for ci, h in enumerate(headers):
        if h in COLUNAS_INTERESSE:
            col_map[h] = ci

    intervalo_col = next((i for i, h in enumerate(headers) if h == "INTERVALO"), 0)

    encontradas = list(col_map.keys())
    ausentes    = [c for c in COLUNAS_INTERESSE if c not in col_map]
    log(f"Colunas encontradas : {encontradas}")
    if ausentes: log(f"Colunas ausentes    : {ausentes}", "AVISO")

    if not encontradas:
        log("Nenhuma coluna de interesse encontrada — colunas no arquivo:", "AVISO")
        log(f"  {headers}")
        return arq

    # Extrair registros
    registros_csv = []
    for row in dados:
        if len(row) <= intervalo_col: continue
        hora = row[intervalo_col].strip()
        if not hora: continue
        for cod in encontradas:
            ci = col_map[cod]
            if ci >= len(row): continue
            raw = row[ci].replace(",", ".").strip()
            try:    mw = float(raw)
            except: mw = 0.0
            ugs = ",".join(MAPA_PDPW.get(cod, [cod]))
            registros_csv.append([data_iso, hora, cod, ugs, mw])

    if not registros_csv:
        log("Nenhum registro extraido", "AVISO")
        return arq

    # Salvar CSV
    arq_csv = arq.parent / f"PDP_ENEVA_N_{data_iso}.csv"
    with open(str(arq_csv), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Data", "Hora", "Codigo_PDPW", "UGs_Dashboard", "MW_Programado"])
        w.writerows(registros_csv)
    log(f"CSV: {arq_csv.name} ({len(registros_csv)} registros)", "OK")

    # Resumo
    log("─" * 55)
    log("RESUMO — MW medio programado:")
    totais: dict[str, list] = {cod: [] for cod in encontradas}
    for rec in registros_csv:
        totais[rec[2]].append(rec[4])
    for cod, vals in totais.items():
        media = sum(vals)/len(vals) if vals else 0
        ugs   = ", ".join(MAPA_PDPW.get(cod, [cod]))
        log(f"  {cod:8s} ({ugs:32s}) {media:7.1f} MW")
    log("─" * 55)

    return arq_csv

# ─── Agendador Windows ────────────────────────────────────────────────────────

def criar_tarefa_agendada():
    script = Path(__file__).resolve()
    python = sys.executable
    cmd = (
        f'schtasks /create /tn "PDP_Download_ENEVA" '
        f'/tr "\\"{python}\\" \\"{script}\\"" '
        f'/sc daily /st 20:30 /ru SYSTEM /f'
    )
    ret = os.system(cmd)
    if ret == 0:
        log("Tarefa criada: PDP_Download_ENEVA, diario 20:30", "OK")
    else:
        log("Falha — execute como Administrador", "ERRO")

# ─── Entry point ─────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Download PDP ENEVA N via PDPW")
    ap.add_argument("--data",    default=None, help="AAAA-MM-DD (padrao: D-1)")
    ap.add_argument("--visible", action="store_true", help="Abrir browser visivel")
    ap.add_argument("--agendar", action="store_true", help="Criar tarefa Windows")
    args = ap.parse_args()

    if args.agendar:
        criar_tarefa_agendada()
        return

    if args.visible:
        CONFIG["headless"] = False

    data_iso = args.data or (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    log("=" * 55)
    log("PDP DOWNLOADER -- Complexo Parnaiba / ENEVA N")
    log(f"Data   : {data_iso}")
    log(f"Destino: {CONFIG['pasta_destino']}")
    log("=" * 55)

    validar_config()
    resultado = baixar_pdp(data_iso)

    if resultado:
        processar_pdp(resultado, data_iso)
        log("=" * 55)
        log(f"SUCESSO: {resultado.name}", "OK")
        log("=" * 55)
        sys.exit(0)
    else:
        log("FALHOU — rode com --visible para depurar", "ERRO")
        sys.exit(1)


if __name__ == "__main__":
    main()
