#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
way2_coleta.py
Busca dados de geração D-1 da API Way2 PIM (5 min) e produz:
  • Way2_UG_5min_{data}.csv  — 5 min, colunas = unit IDs do dashboard (eventos)
  • Way2_UG_30min_{data}.csv — 30 min agregado (média 6 slots), para geração
  • Way2_D1_{data}.xlsx      — planilha para avaliação humana

Uso:  python way2_coleta.py
"""

import sys, io
import requests
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Configuração ───────────────────────────────────────────────────────────────
WAY2_BASE  = "https://pim.way2.com.br:183/api/v3/dados-de-medicao/pontos"
WAY2_TOKEN = "3f503c11-24db-4e46-aed0-831fc3bd7567"
HEADERS    = {"pim-auth": WAY2_TOKEN}
DEV_DIR    = Path(__file__).parent

# ── Mapeamento Way2 → unit IDs do dashboard (VERIFICAR com operação se necessário) ──
WAY2_TO_UNIT = {
    # P2 - Maranhão III
    "UTEMaranhao3_UG51B":    "TG51",
    "UTEMaranhao3_UG52B":    "TG52",
    "UTEMaranhao3_UG58B":    "TV58",
    # P3-6 - Nova Venécia
    "UTE Nova Venécia UG18": "TV18",
    "UTENovaVenécia_UG12":   "TG12",
    "UTENovaVenécia_UG10":   "UG10",
    # P4 - Parnaíba IV
    "UTE Parnaiba IV ML01":  "BAG011",
    "UTE Parnaiba IV ML02":  "BAG021",
    "UTE Parnaiba IV ML03":  "BAG031",
    # P1-5 - Maranhão IV → TG31/TG32 | Maranhão V → TG21/TG22 | Parnaíba V → TV28
    "UTEMaranhao4_UG1":      "TG31",
    "UTEMaranhao4_UG2":      "TG32",
    "UTEMaranhao5_UG1":      "TG21",
    "UTEMaranhao5_UG2":      "TG22",
    "SEParnaibaV_UG1":       "TV28",
}

# Ordem das colunas alinhada com UNITS do dashboard
UNIT_ORDER = ["TG31","TG32","TG21","TG22","TV28",
              "TG51","TG52","TV58","TG12","TV18","UG10",
              "BAG011","BAG021","BAG031"]

# ── Grupos: nome → {ids, mapa ID→rótulo} ──────────────────────────────────────
GRUPOS = {
    "P2 - Maranhão III": {
        "ids": [760, 464, 700, 463, 488, 462, 1116, 894],
        "mapa": {
            "760":  "UTEMaranhao3_UG58B",
            "464":  "SEParnaiba_TR58",
            "700":  "UTEMaranhao3_UG52B",
            "463":  "SEParnaiba_TR52",
            "488":  "UTEMaranhao3_UG51B",
            "462":  "SEParnaiba_TR51",
            "1116": "MaranhãoIII Bruta",
            "894":  "ParnaíbaBruta",
        },
    },
    "P3-6 - Nova Venécia": {
        "ids": [6113, 5945, 433, 546, 1122, 894],
        "mapa": {
            "6113": "UTE Nova Venécia UG18",
            "5945": "Nova Venecia II Bruta",
            "433":  "UTENovaVenécia_UG12",
            "546":  "UTENovaVenécia_UG10",
            "1122": "Nova Venécia Bruta",
            "894":  "ParnaíbaBruta",
        },
    },
    "P4 - Parnaíba IV": {
        "ids": [5903, 5904, 5905, 1123, 894],
        "mapa": {
            "5903": "UTE Parnaiba IV ML01",
            "5904": "UTE Parnaiba IV ML02",
            "5905": "UTE Parnaiba IV ML03",
            "1123": "Parnaíba IV Bruta",
            "894":  "ParnaíbaBruta",
        },
    },
    "P1-5 - Maranhão IV/V + ParnV": {
        "ids": [377, 373, 375, 374, 404, 371, 379, 372,
                1116, 1118, 1120, 1122, 1123, 3528, 3527],
        "mapa": {
            "377":  "UTEMaranhao4_UG2",
            "373":  "SEParnaiba_TR13",
            "375":  "UTEMaranhao4_UG1",
            "374":  "SEParnaiba_TR12",
            "404":  "UTEMaranhao5_UG2",
            "371":  "SEParnaiba_TR11",
            "379":  "UTEMaranhao5_UG1",
            "372":  "SEParnaiba_TR10",
            "1116": "MaranhãoIII Bruta",
            "1118": "Maranhão IV Bruta",
            "1120": "Maranhão V Bruta",
            "1122": "Nova Venécia Bruta",
            "1123": "Parnaíba IV Bruta",
            "3528": "SEParnaibaV_UG1",
            "3527": "SEParnaíbaV_TR28",
        },
    },
}


def _parse_dt(s: str) -> datetime | None:
    if not s:
        return None
    s_clean = s[:19].replace("T", " ")
    try:
        return datetime.strptime(s_clean, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def buscar_grupo(nome: str, cfg: dict) -> pd.DataFrame:
    """Chama a API Way2 (5 min) e retorna DataFrame pivotado com valores em MW."""
    ids_str = "%2C".join(str(i) for i in cfg["ids"])
    url = (
        f"{WAY2_BASE}?ids={ids_str}"
        "&grandezas=Eneat"
        "&contextodasdatas=ConsiderarDiaCheio"
        "&intervalo=CincoMinutos"
        "&medicao-temporelativo=Ontem"
    )

    print(f"  [{nome}] requisitando...")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=60, verify=False)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  ERRO: {e}")
        return pd.DataFrame()

    payload = resp.json()

    root_values = list(payload.values())
    points_list = None
    if len(root_values) > 5 and isinstance(root_values[5], list):
        points_list = root_values[5]
    else:
        for v in root_values:
            if isinstance(v, list) and v and isinstance(v[0], dict) and "pontoId" in v[0]:
                points_list = v
                break

    if not points_list:
        print(f"  AVISO: estrutura JSON inesperada.")
        return pd.DataFrame()

    rows = []
    for ponto in points_list:
        pid = str(ponto.get("pontoId", ""))
        for v in ponto.get("valores", []):
            dt = _parse_dt(v.get("data", ""))
            if dt is None:
                continue
            # Arredonda para baixo ao bloco de 5 min
            bloco = (dt.minute // 5) * 5
            datahora = dt.replace(minute=bloco, second=0, microsecond=0)
            rows.append({"DataHora": datahora, "PontoId": pid, "kWh": v.get("valor")})

    if not rows:
        print(f"  AVISO: nenhum valor retornado.")
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df_agg = (
        df.dropna(subset=["kWh"])
          .groupby(["DataHora", "PontoId"])["kWh"]
          .max()
          .reset_index()
    )

    df_pivot = df_agg.pivot(index="DataHora", columns="PontoId", values="kWh")
    df_pivot.columns.name = None
    df_pivot.sort_index(inplace=True)

    mapa = cfg["mapa"]
    df_pivot.rename(columns={k: v for k, v in mapa.items()}, inplace=True)

    # kWh → MW  (5 min = 5/60 h  →  MW = kWh × 60/5 / 1000 = kWh × 12 / 1000)
    for col in df_pivot.columns:
        df_pivot[col] = (df_pivot[col] * 12 / 1000.0).round(3)

    df_pivot.reset_index(inplace=True)
    print(f"  OK: {len(df_pivot)} slots  |  {len(df_pivot.columns) - 1} pontos")
    return df_pivot


def _eh_ug(col: str) -> bool:
    c = col.upper()
    return "BRUTA" not in c and "_TR" not in c


def _combinar_ugs(dfs: dict) -> pd.DataFrame | None:
    """Junta todos os grupos, retendo apenas colunas de UG (sem TR, sem Bruta)."""
    combined: pd.DataFrame | None = None
    seen: set[str] = set()
    for df in dfs.values():
        if df.empty:
            continue
        ug_cols = ["DataHora"] + [c for c in df.columns if c != "DataHora" and _eh_ug(c)]
        df_c = df[ug_cols].copy()
        novas = [c for c in df_c.columns if c == "DataHora" or c not in seen]
        df_c = df_c[novas]
        seen.update(c for c in novas if c != "DataHora")
        combined = df_c if combined is None else combined.merge(df_c, on="DataHora", how="outer")
    if combined is not None:
        combined.sort_values("DataHora", inplace=True)
        combined.reset_index(drop=True, inplace=True)
        # API usa fim-de-slot: o último slot (23:55-00:00) chega como 00:00 do dia seguinte.
        # Converte para início do slot (−5 min) para manter tudo no dia alvo.
        mask = combined["DataHora"].dt.hour == 0
        mask &= combined["DataHora"].dt.minute == 0
        mask &= combined["DataHora"].dt.date != combined["DataHora"].iloc[0].date()
        if mask.any():
            combined.loc[mask, "DataHora"] -= pd.Timedelta(minutes=5)
    return combined


def gerar_csvs_dashboard(dfs: dict, d1) -> None:
    """
    Gera Way2_UG_5min_{d1}.csv e Way2_UG_30min_{d1}.csv para o dashboard.
    Colunas = unit IDs do dashboard (TG31, TG32, …).
    """
    combined = _combinar_ugs(dfs)
    if combined is None or combined.empty:
        print("  AVISO: nenhum dado para gerar CSVs.")
        return

    # Renomeia Way2 → unit IDs
    combined.rename(columns=WAY2_TO_UNIT, inplace=True)

    # Ordena colunas
    unit_cols = [c for c in UNIT_ORDER if c in combined.columns]
    extras    = [c for c in combined.columns if c not in UNIT_ORDER and c != "DataHora"]
    combined  = combined[["DataHora"] + unit_cols + extras]

    # CSV 5 min
    arq5 = DEV_DIR / f"Way2_UG_5min_{d1}.csv"
    combined.to_csv(str(arq5), index=False)
    print(f"  CSV 5 min  → {arq5.name}  ({len(combined)} linhas)")

    # CSV 30 min  (média de 6 slots de 5 min por janela de 30 min)
    combined_30 = combined.copy()
    combined_30["DataHora"] = combined_30["DataHora"].dt.floor("30min")
    combined_30 = (
        combined_30
        .groupby("DataHora")[unit_cols + extras]
        .mean()
        .round(3)
        .reset_index()
    )
    arq30 = DEV_DIR / f"Way2_UG_30min_{d1}.csv"
    combined_30.to_csv(str(arq30), index=False)
    print(f"  CSV 30 min → {arq30.name}  ({len(combined_30)} linhas)")


def _estilizar(ws, ncols: int):
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 20
    for i in range(2, ncols + 2):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = 24
    ws.freeze_panes = "B2"


def main():
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    d1 = (datetime.now() - timedelta(days=1)).date()
    print(f"\n=== Way2 Coleta D-1: {d1} ===\n")

    dfs: dict[str, pd.DataFrame] = {}
    for nome, cfg in GRUPOS.items():
        dfs[nome] = buscar_grupo(nome, cfg)

    # ── CSVs para o dashboard ──────────────────────────────────────────────────
    print("\nGerando CSVs para o dashboard...")
    gerar_csvs_dashboard(dfs, d1)

    # ── XLSX para avaliação (Combinado 5 min + abas individuais) ───────────────
    saida = DEV_DIR / f"Way2_D1_{d1}.xlsx"

    with pd.ExcelWriter(saida, engine="openpyxl") as writer:

        # Aba Combinado: unit IDs, 5 min
        combined_xlsx = _combinar_ugs(dfs)
        if combined_xlsx is not None:
            combined_xlsx.rename(columns=WAY2_TO_UNIT, inplace=True)
            unit_cols_x = [c for c in UNIT_ORDER if c in combined_xlsx.columns]
            extras_x    = [c for c in combined_xlsx.columns
                           if c not in UNIT_ORDER and c != "DataHora"]
            combined_xlsx = combined_xlsx[["DataHora"] + unit_cols_x + extras_x]
            combined_xlsx.to_excel(writer, sheet_name="Combinado", index=False)
            _estilizar(writer.sheets["Combinado"], len(combined_xlsx.columns) - 1)

        # Abas individuais (nomes Way2 originais, todos os pontos)
        for nome, df in dfs.items():
            sheet = nome.replace("/", "-")[:31]
            if df.empty:
                pd.DataFrame({"Sem dados": []}).to_excel(writer, sheet_name=sheet, index=False)
            else:
                df.to_excel(writer, sheet_name=sheet, index=False)
                _estilizar(writer.sheets[sheet], len(df.columns) - 1)

    print(f"\nPlanilha salva em:\n  {saida}\n")


if __name__ == "__main__":
    main()
